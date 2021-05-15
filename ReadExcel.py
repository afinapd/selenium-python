import openpyxl

path="D:\\Timesheet SMM\\1. Afina Putri Dayanti - Mei.xlsx"

workbook = openpyxl.load_workbook(path)

sheet= workbook.active #workbook.get_sheet_by_name("Name")
rows = sheet.max_row
cols = sheet.max_column

print(rows) #48
print(cols) #10

for row in range(5,rows+1) :
    for col in range (2,cols+1) :
        print(sheet.cell(row=row, column=col).value, end="       ")
    print()

